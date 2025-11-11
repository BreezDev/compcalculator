import io
import json
from datetime import date, datetime

from dateutil.relativedelta import relativedelta
from flask import (
    Flask,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from flask_login import (
    LoginManager,
    UserMixin,
    current_user,
    login_required,
    login_user,
    logout_user,
)
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func
from werkzeug.security import check_password_hash, generate_password_hash
from openpyxl.styles import Font, PatternFill
import pandas as pd


DEFAULT_PASSWORD = "MisoMochi722!"
P_AND_C_CATEGORIES = {
    "auto",
    "homeowners",
    "renters",
    "plup",
    "pap",
    "other_fire",
    "bank",
    "business",
}
FS_CATEGORIES = {"life", "health"}

CATEGORY_LABELS = [
    ("auto", "Auto"),
    ("homeowners", "Homeowners"),
    ("renters", "Renters"),
    ("plup", "Personal Umbrella"),
    ("pap", "Personal Articles"),
    ("other_fire", "Other Fire"),
    ("life", "Life"),
    ("health", "Health"),
    ("bank", "Bank"),
    ("business", "Business"),
]

app = Flask(__name__)
app.config["SECRET_KEY"] = "change-this-secret-key"
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///commission_tracker.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = "login"


class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    sales = db.relationship("Sale", back_populates="user", cascade="all, delete-orphan")

    def set_password(self, password: str) -> None:
        self.password_hash = generate_password_hash(password)

    def check_password(self, password: str) -> bool:
        return check_password_hash(self.password_hash, password)


class Sale(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    client_name = db.Column(db.String(120), nullable=False)
    date_sold = db.Column(db.Date, nullable=False)
    date_effective = db.Column(db.Date, nullable=False)
    category = db.Column(db.String(40), nullable=False)
    premium = db.Column(db.Float, nullable=False)
    fs_monthly_premium = db.Column(db.Float)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    user = db.relationship("User", back_populates="sales")


@login_manager.user_loader
def load_user(user_id: str):
    return db.session.get(User, int(user_id))


def ensure_database() -> None:
    with app.app_context():
        db.create_all()

@app.context_processor
def inject_current_year():
    return {"current_year": datetime.utcnow().year}


@app.route("/")
def landing():
    return render_template("landing.html")


@app.route("/signup", methods=["GET", "POST"])
def signup():
    if request.method == "POST":
        username = request.form.get("username", "").strip().lower()
        confirmation = request.form.get("confirmation", "")

        if not username:
            flash("Please provide a username.", "danger")
            return redirect(url_for("signup"))

        if confirmation != DEFAULT_PASSWORD:
            flash("Please confirm the default team password correctly.", "danger")
            return redirect(url_for("signup"))

        existing_user = User.query.filter(func.lower(User.username) == username).first()
        if existing_user:
            flash("That username is already taken.", "warning")
            return redirect(url_for("signup"))

        user = User(username=username)
        user.set_password(DEFAULT_PASSWORD)
        db.session.add(user)
        db.session.commit()
        flash("Account created! You can now log in.", "success")
        return redirect(url_for("login"))

    return render_template("signup.html", default_password=DEFAULT_PASSWORD)


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip().lower()
        password = request.form.get("password", "")

        user = User.query.filter(func.lower(User.username) == username).first()
        if user and user.check_password(password):
            login_user(user)
            flash("Welcome back!", "success")
            next_page = request.args.get("next")
            return redirect(next_page or url_for("enter_sale"))

        flash("Invalid credentials. Please try again.", "danger")
    return render_template("login.html", default_password=DEFAULT_PASSWORD)


@app.route("/logout")
@login_required
def logout():
    logout_user()
    flash("You have been logged out.", "info")
    return redirect(url_for("landing"))


@app.route("/enter-sale", methods=["GET", "POST"])
@login_required
def enter_sale():
    if request.method == "POST":
        client_name = request.form.get("client_name", "").strip()
        date_sold_raw = request.form.get("date_sold")
        date_effective_raw = request.form.get("date_effective")
        category = request.form.get("category")
        premium_raw = request.form.get("premium")
        fs_premium_raw = request.form.get("fs_monthly_premium")

        if not client_name or not date_sold_raw or not date_effective_raw or not premium_raw:
            flash("All fields except FS Monthly Premium are required.", "danger")
            return redirect(url_for("enter_sale"))

        try:
            date_sold = datetime.strptime(date_sold_raw, "%Y-%m-%d").date()
            date_effective = datetime.strptime(date_effective_raw, "%Y-%m-%d").date()
            premium = float(premium_raw)
            fs_monthly_premium = float(fs_premium_raw) if fs_premium_raw else None
        except ValueError:
            flash("Please enter valid numbers and dates.", "danger")
            return redirect(url_for("enter_sale"))

        if category not in dict(CATEGORY_LABELS):
            flash("Please choose a valid category.", "danger")
            return redirect(url_for("enter_sale"))

        sale = Sale(
            user=current_user,
            client_name=client_name,
            date_sold=date_sold,
            date_effective=date_effective,
            category=category,
            premium=premium,
            fs_monthly_premium=fs_monthly_premium,
        )
        db.session.add(sale)
        db.session.commit()
        flash("Sale recorded successfully!", "success")
        return redirect(url_for("my_sales"))

    return render_template("enter_sale.html", categories=CATEGORY_LABELS)


@app.route("/my-sales")
@login_required
def my_sales():
    sales = (
        Sale.query.filter_by(user_id=current_user.id)
        .order_by(Sale.date_sold.desc())
        .all()
    )
    return render_template("my_sales.html", sales=sales, categories=dict(CATEGORY_LABELS))


def parse_date(date_str: str | None) -> date | None:
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        return None


@app.route("/team-sales")
@login_required
def team_sales():
    start_date = parse_date(request.args.get("start_date"))
    end_date = parse_date(request.args.get("end_date"))

    query = Sale.query.join(User)
    if start_date:
        query = query.filter(Sale.date_sold >= start_date)
    if end_date:
        query = query.filter(Sale.date_sold <= end_date)

    sales = query.order_by(Sale.date_sold.desc()).all()

    category_labels_map = dict(CATEGORY_LABELS)
    category_totals = {label: 0 for _, label in CATEGORY_LABELS}
    monthly_totals: dict[date, float] = {}

    for sale in sales:
        category_label = category_labels_map[sale.category]
        category_totals[category_label] += sale.premium
        month_key = sale.date_sold.replace(day=1)
        monthly_totals.setdefault(month_key, 0)
        monthly_totals[month_key] += sale.premium

    sorted_months = sorted(monthly_totals.keys())

    chart_data = {
        "category_labels": list(category_totals.keys()),
        "category_values": [round(value, 2) for value in category_totals.values()],
        "monthly_labels": [month.strftime("%b %Y") for month in sorted_months],
        "monthly_values": [round(monthly_totals[month], 2) for month in sorted_months],
    }

    return render_template(
        "team_sales.html",
        sales=sales,
        start_date=start_date,
        end_date=end_date,
        categories=dict(CATEGORY_LABELS),
        chart_data=json.dumps(chart_data),
    )


@app.route("/team-sales/export")
@login_required
def export_team_sales():
    start_date = parse_date(request.args.get("start_date"))
    end_date = parse_date(request.args.get("end_date"))

    query = Sale.query.join(User)
    if start_date:
        query = query.filter(Sale.date_sold >= start_date)
    if end_date:
        query = query.filter(Sale.date_sold <= end_date)

    sales = query.order_by(Sale.date_sold.asc()).all()

    if not sales:
        flash("No sales found for the selected period.", "warning")
        return redirect(url_for("team_sales"))

    records = []
    for sale in sales:
        records.append(
            {
                "Client Name": sale.client_name,
                "Team Member": sale.user.username,
                "Date Sold": sale.date_sold.strftime("%Y-%m-%d"),
                "Date Effective": sale.date_effective.strftime("%Y-%m-%d"),
                "Category": dict(CATEGORY_LABELS)[sale.category],
                "Policy Premium": sale.premium,
                "FS Monthly Premium": sale.fs_monthly_premium or 0,
            }
        )

    df = pd.DataFrame(records)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Team Sales")
        worksheet = writer.sheets["Team Sales"]
        header_font = Font(bold=True, color="FFFFFF")
        fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = fill
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length + 4

    output.seek(0)
    filename_parts = ["team-sales"]
    if start_date:
        filename_parts.append(start_date.strftime("%Y%m%d"))
    if end_date:
        filename_parts.append(end_date.strftime("%Y%m%d"))
    filename = "-".join(filename_parts) + ".xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/my-sales/export")
@login_required
def export_my_sales():
    sales = (
        Sale.query.filter_by(user_id=current_user.id)
        .order_by(Sale.date_sold.asc())
        .all()
    )

    if not sales:
        flash("You do not have any sales to export yet.", "warning")
        return redirect(url_for("my_sales"))

    records = []
    for sale in sales:
        records.append(
            {
                "Client Name": sale.client_name,
                "Date Sold": sale.date_sold.strftime("%Y-%m-%d"),
                "Date Effective": sale.date_effective.strftime("%Y-%m-%d"),
                "Category": dict(CATEGORY_LABELS)[sale.category],
                "Policy Premium": sale.premium,
                "FS Monthly Premium": sale.fs_monthly_premium or 0,
            }
        )

    df = pd.DataFrame(records)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="My Sales")
        worksheet = writer.sheets["My Sales"]
        header_font = Font(bold=True, color="FFFFFF")
        fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = fill
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length + 4

    output.seek(0)
    filename = f"{current_user.username}-sales.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def get_commission_rate(p_c_apps: int, life_apps: int) -> float:
    if p_c_apps <= 0:
        return 0.0

    if life_apps >= 5:
        brackets = {"1-19": 0.07, "20-29": 0.09, "30-39": 0.10, "40-49": 0.11, "50+": 0.12}
    elif life_apps == 4:
        brackets = {"1-19": 0.06, "20-29": 0.08, "30-39": 0.09, "40-49": 0.10, "50+": 0.11}
    elif life_apps == 3:
        brackets = {"1-19": 0.05, "20-29": 0.07, "30-39": 0.08, "40-49": 0.09, "50+": 0.10}
    elif life_apps == 2:
        brackets = {"1-19": 0.04, "20-29": 0.06, "30-39": 0.07, "40-49": 0.08, "50+": 0.09}
    else:
        brackets = {"1-19": 0.02, "20-29": 0.04, "30-39": 0.05, "40-49": 0.06, "50+": 0.07}

    if p_c_apps >= 50:
        return brackets["50+"]
    if p_c_apps >= 40:
        return brackets["40-49"]
    if p_c_apps >= 30:
        return brackets["30-39"]
    if p_c_apps >= 20:
        return brackets["20-29"]
    if p_c_apps >= 1:
        return brackets["1-19"]
    return 0.0


def fs_bonus(fs_monthly_premium: float) -> float:
    if fs_monthly_premium >= 500:
        return 800.0
    if fs_monthly_premium >= 400:
        return 600.0
    if fs_monthly_premium >= 300:
        return 400.0
    if fs_monthly_premium >= 200:
        return 200.0
    return 0.0


def life_app_bonus(life_apps: int) -> float:
    if life_apps < 6:
        return 0.0
    bonus = 500.0
    if life_apps > 6:
        bonus += (life_apps - 6) * 150.0
    return bonus


def milestone_bonus(p_c_apps: int, life_apps: int) -> float:
    if p_c_apps >= 30 and life_apps >= 4:
        return 1000.0
    if p_c_apps >= 25 and life_apps >= 3:
        return 750.0
    if p_c_apps >= 20 and life_apps >= 2:
        return 500.0
    return 0.0


@app.route("/commission-calculator", methods=["GET", "POST"])
@login_required
def commission_calculator():
    target_month_str = request.values.get("target_month")
    if target_month_str:
        try:
            target_month = datetime.strptime(target_month_str, "%Y-%m")
        except ValueError:
            flash("Please select a valid month.", "danger")
            return redirect(url_for("commission_calculator"))
    else:
        target_month = datetime.utcnow()
        target_month_str = target_month.strftime("%Y-%m")

    start_of_month = date(target_month.year, target_month.month, 1)
    end_of_month = start_of_month + relativedelta(months=1) - relativedelta(days=1)

    sales = (
        Sale.query.filter_by(user_id=current_user.id)
        .filter(Sale.date_sold >= start_of_month, Sale.date_sold <= end_of_month)
        .all()
    )

    p_c_premium = sum(sale.premium for sale in sales if sale.category in P_AND_C_CATEGORIES)
    p_c_apps = sum(1 for sale in sales if sale.category in P_AND_C_CATEGORIES)
    life_apps = sum(1 for sale in sales if sale.category == "life")
    fs_monthly_total = sum(
        sale.fs_monthly_premium or 0 for sale in sales if sale.category in FS_CATEGORIES
    )

    commission_rate = get_commission_rate(p_c_apps, life_apps)
    commission_amount = p_c_premium * commission_rate if p_c_premium >= 12000 else 0.0
    commission_eligible = p_c_premium >= 12000

    fs_monthly_bonus = fs_bonus(fs_monthly_total)
    fs_high_bonus = 1000.0 if fs_monthly_total > 1000 else 0.0
    life_bonus_amount = life_app_bonus(life_apps)
    milestone_bonus_amount = milestone_bonus(p_c_apps, life_apps)

    total_bonus = fs_monthly_bonus + fs_high_bonus + life_bonus_amount + milestone_bonus_amount
    total_compensation = commission_amount + total_bonus

    context = {
        "target_month": target_month_str,
        "sales": sales,
        "p_c_premium": p_c_premium,
        "p_c_apps": p_c_apps,
        "life_apps": life_apps,
        "fs_monthly_total": fs_monthly_total,
        "commission_rate": commission_rate,
        "commission_amount": commission_amount,
        "commission_eligible": commission_eligible,
        "fs_monthly_bonus": fs_monthly_bonus,
        "fs_high_bonus": fs_high_bonus,
        "life_bonus_amount": life_bonus_amount,
        "milestone_bonus_amount": milestone_bonus_amount,
        "total_bonus": total_bonus,
        "total_compensation": total_compensation,
    }
    return render_template("commission_calculator.html", **context, categories=dict(CATEGORY_LABELS))


ensure_database()


if __name__ == "__main__":
    app.run(debug=True)
